import os
from openpyxl import load_workbook
import sys
print(sys.executable)

# 🔧 SETTINGS – edit these before running
root_folder = r"C:\Users\judep\Downloads\DDD"  # Folder containing .xlsx files
find_text = "Belships"
replace_text = "Global Maritime Ship Management, Inc. (GMSMI)"

# --- Do not edit below this line ---
count_files = 0
count_replaced = 0

for foldername, subfolders, filenames in os.walk(root_folder):
    for filename in filenames:
        if filename.endswith(".xlsx") and not filename.startswith("~$"):  # skip temp Excel files
            file_path = os.path.join(foldername, filename)
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"⚠️ Skipped {filename} (error reading file: {e})")
                continue

            replaced_in_file = False

            # Loop through all sheets and cells
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and find_text in cell.value:
                            cell.value = cell.value.replace(find_text, replace_text)
                            replaced_in_file = True

            if replaced_in_file:
                wb.save(file_path)
                count_replaced += 1
                print(f"✅ Modified: {file_path}")
            else:
                print(f"— No change: {file_path}")

            count_files += 1

print(f"\n✅ Processed {count_files} Excel files (including subfolders).")
print(f"🔄 Updated {count_replaced} files containing '{find_text}'.")
print("🎉 Done!")
